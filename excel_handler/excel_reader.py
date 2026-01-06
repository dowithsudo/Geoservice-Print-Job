from openpyxl import load_workbook
from datetime import datetime

REQUIRED_HEADERS = [
    "job_no",
    "box_no",
    "first",
    "last",
    "date_received",
    "lsn",
    "sid",
]


def read_excel(file_path):
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("File Excel kosong")

    headers = [str(h).strip() if h else "" for h in rows[0]]

    if headers != REQUIRED_HEADERS:
        raise ValueError(
            f"Header Excel tidak sesuai.\n"
            f"Harus: {REQUIRED_HEADERS}\n"
            f"Dapat: {headers}"
        )

    data_list = []

    for row in rows[1:]:
        # Skip baris yang semua kolomnya kosong
        if row is None or all(
            cell is None or str(cell).strip() == ""
            for cell in row
        ):
            continue

        data = {}

        for key, value in zip(REQUIRED_HEADERS, row):
            if value is None:
                data[key] = ""
                continue

            # KHUSUS DATE
            if key == "date_received":
                if isinstance(value, datetime):
                    # Excel datetime → dd/mm/yy
                    data[key] = value.strftime("%d/%m/%y")
                else:
                    # Kalau sudah string, buang jam kalau ada
                    date_str = str(value).split(" ")[0]

                    # Kalau format yyyy-mm-dd → ubah ke dd/mm/yy
                    try:
                        parsed = datetime.strptime(date_str, "%Y-%m-%d")
                        data[key] = parsed.strftime("%d/%m/%y")
                    except ValueError:
                        # Anggap sudah benar (misal 30/12/25)
                        data[key] = date_str
            else:
                data[key] = str(value)

        data_list.append(data)

    return data_list
